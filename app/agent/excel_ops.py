"""Excel file operations — read, summarize, and apply changes."""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def summarize_excel(path: str, preview_rows: int = 5) -> dict:
    """Read an Excel file and return a structured summary.

    Returns:
        {
            "filename": "...",
            "sheets": [{"name": "...", "rows": N, "cols": N, "headers": [...]}],
            "sheet_summary": "...",
            "data_preview": "...",
        }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets_info = []
    summary_parts = []
    preview_parts = []

    for ws in wb.worksheets:
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        headers = []
        if rows > 0:
            headers = [str(ws.cell(1, c).value or "") for c in range(1, cols + 1)]

        sheets_info.append({
            "name": ws.title,
            "rows": rows,
            "cols": cols,
            "headers": headers,
        })

        summary_parts.append(
            f"- *{ws.title}*: {rows} rows x {cols} cols | Headers: {', '.join(headers[:10])}"
            + (" ..." if len(headers) > 10 else "")
        )

        # Data preview
        preview_lines = [f"**{ws.title}**"]
        for r in range(1, min(rows + 1, preview_rows + 2)):  # +1 for header, +1 for range
            row_data = []
            for c in range(1, min(cols + 1, 10)):  # cap at 10 cols for readability
                val = ws.cell(r, c).value
                row_data.append(str(val) if val is not None else "")
            preview_lines.append(" | ".join(row_data))
        preview_parts.append("\n".join(preview_lines))

    wb.close()

    return {
        "filename": Path(path).name,
        "sheets": sheets_info,
        "sheet_summary": "\n".join(summary_parts),
        "data_preview": "\n\n".join(preview_parts),
    }


def apply_actions(path: str, actions: list[dict]) -> str:
    """Apply a list of actions to an Excel file. Returns path to the modified file.

    Supported actions:
        - update_cell: {"sheet": "...", "cell": "A1", "value": "..."}
        - add_row: {"sheet": "...", "values": [...]}
        - add_sheet: {"name": "..."}
    """
    wb = openpyxl.load_workbook(path)
    changes = []

    for action in actions:
        action_type = action.get("type")

        if action_type == "update_cell":
            sheet_name = action["sheet"]
            ws = wb[sheet_name]
            cell = action["cell"]
            value = _coerce_value(action["value"])
            old_value = ws[cell].value
            ws[cell] = value
            changes.append(f"Updated {sheet_name}!{cell}: {old_value} -> {value}")

        elif action_type == "add_row":
            sheet_name = action["sheet"]
            ws = wb[sheet_name]
            values = [_coerce_value(v) for v in action["values"]]
            ws.append(values)
            changes.append(f"Added row to {sheet_name}: {values}")

        elif action_type == "add_sheet":
            name = action["name"]
            wb.create_sheet(title=name)
            changes.append(f"Created new sheet: {name}")

        else:
            logger.warning("Unknown action type: %s", action_type)

    # Save to same path (overwrite)
    wb.save(path)
    wb.close()

    logger.info("Applied %d actions: %s", len(changes), changes)
    return path


def _coerce_value(val: Any) -> Any:
    """Try to convert string values to appropriate types."""
    if not isinstance(val, str):
        return val
    # Try numeric
    try:
        if "." in val:
            return float(val)
        return int(val)
    except (ValueError, TypeError):
        pass
    return val
